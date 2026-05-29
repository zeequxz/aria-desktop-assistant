"""
agent/file_tools.py - File system tools available to the AI agent.

Each tool is a plain Python function. The agent decides when to call them.
All destructive operations describe what they'll do before doing it.
"""

import os
import shutil
import json
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Optional


# ── Directory listing ──────────────────────────────────────────────────────

def list_directory(path: str, recursive: bool = False) -> dict:
    """List files and folders at a given path."""
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"Path does not exist: {path}"}
        if not p.is_dir():
            return {"error": f"Not a directory: {path}"}

        entries = []
        if recursive:
            for item in sorted(p.rglob("*"))[:200]:  # cap at 200
                entries.append({
                    "name": item.name,
                    "path": str(item),
                    "type": "folder" if item.is_dir() else "file",
                    "size_kb": round(item.stat().st_size / 1024, 1) if item.is_file() else None,
                    "modified": datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
        else:
            for item in sorted(p.iterdir()):
                entries.append({
                    "name": item.name,
                    "path": str(item),
                    "type": "folder" if item.is_dir() else "file",
                    "size_kb": round(item.stat().st_size / 1024, 1) if item.is_file() else None,
                    "modified": datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
        return {"path": str(p), "entries": entries, "count": len(entries)}
    except PermissionError:
        return {"error": f"Permission denied: {path}"}
    except Exception as e:
        return {"error": str(e)}


# ── Read file ──────────────────────────────────────────────────────────────

def read_file(path: str, max_chars: int = 8000) -> dict:
    """Read the contents of a text file."""
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"File not found: {path}"}
        if not p.is_file():
            return {"error": f"Not a file: {path}"}

        suffix = p.suffix.lower()

        # Plain text / code / config
        if suffix in {".txt", ".md", ".py", ".js", ".ts", ".json", ".csv",
                      ".xml", ".html", ".htm", ".css", ".bat", ".sh", ".log",
                      ".ini", ".cfg", ".yaml", ".yml", ".toml", ".sql"}:
            content = p.read_text(encoding="utf-8", errors="replace")
            truncated = len(content) > max_chars
            return {
                "path": str(p),
                "content": content[:max_chars],
                "truncated": truncated,
                "size_kb": round(p.stat().st_size / 1024, 1),
            }

        # Word documents
        if suffix == ".docx":
            try:
                from docx import Document
                doc = Document(str(p))
                text = "\n".join(para.text for para in doc.paragraphs)
                truncated = len(text) > max_chars
                return {"path": str(p), "content": text[:max_chars], "truncated": truncated, "type": "docx"}
            except ImportError:
                return {"error": "python-docx not installed"}

        # Excel files
        if suffix in {".xlsx", ".xls"}:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    lines.append(f"=== Sheet: {sheet} ===")
                    for row in ws.iter_rows(max_row=100, values_only=True):
                        lines.append("\t".join(str(c or "") for c in row))
                text = "\n".join(lines)
                truncated = len(text) > max_chars
                return {"path": str(p), "content": text[:max_chars], "truncated": truncated, "type": "excel"}
            except ImportError:
                return {"error": "openpyxl not installed"}

        return {"error": f"Unsupported file type: {suffix}. Supported: txt, md, py, js, json, csv, docx, xlsx and more."}
    except Exception as e:
        return {"error": str(e)}


# ── Write file ─────────────────────────────────────────────────────────────

def write_file(path: str, content: str, overwrite: bool = False) -> dict:
    """Write content to a file. Creates parent directories if needed."""
    try:
        p = Path(path).expanduser().resolve()
        if p.exists() and not overwrite:
            return {"error": f"File already exists: {path}. Set overwrite=True to replace it."}
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")
        return {"success": True, "path": str(p), "size_kb": round(p.stat().st_size / 1024, 1)}
    except PermissionError:
        return {"error": f"Permission denied writing to: {path}"}
    except Exception as e:
        return {"error": str(e)}


# ── Copy / move / delete ───────────────────────────────────────────────────

def copy_file(src: str, dst: str) -> dict:
    try:
        s = Path(src).expanduser().resolve()
        d = Path(dst).expanduser().resolve()
        if not s.exists():
            return {"error": f"Source not found: {src}"}
        d.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(s), str(d))
        return {"success": True, "copied_to": str(d)}
    except Exception as e:
        return {"error": str(e)}


def move_file(src: str, dst: str) -> dict:
    try:
        s = Path(src).expanduser().resolve()
        d = Path(dst).expanduser().resolve()
        if not s.exists():
            return {"error": f"Source not found: {src}"}
        d.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(s), str(d))
        return {"success": True, "moved_to": str(d)}
    except Exception as e:
        return {"error": str(e)}


def delete_file(path: str, send_to_recycle: bool = True) -> dict:
    """Delete a file. On Windows, sends to Recycle Bin by default."""
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"File not found: {path}"}
        if send_to_recycle and os.name == "nt":
            try:
                import winreg  # noqa - just checking we're on Windows
                subprocess.run(
                    ["powershell", "-Command",
                     f"Add-Type -AssemblyName Microsoft.VisualBasic; [Microsoft.VisualBasic.FileIO.FileSystem]::DeleteFile('{str(p)}','OnlyErrorDialogs','SendToRecycleBin')"],
                    capture_output=True
                )
                return {"success": True, "sent_to_recycle_bin": str(p)}
            except Exception:
                pass
        if p.is_dir():
            shutil.rmtree(str(p))
        else:
            p.unlink()
        return {"success": True, "deleted": str(p)}
    except Exception as e:
        return {"error": str(e)}


def create_folder(path: str) -> dict:
    try:
        p = Path(path).expanduser().resolve()
        p.mkdir(parents=True, exist_ok=True)
        return {"success": True, "created": str(p)}
    except Exception as e:
        return {"error": str(e)}


def rename_file(path: str, new_name: str) -> dict:
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"Not found: {path}"}
        new_path = p.parent / new_name
        p.rename(new_path)
        return {"success": True, "renamed_to": str(new_path)}
    except Exception as e:
        return {"error": str(e)}


def search_files(folder: str, pattern: str, content_search: Optional[str] = None) -> dict:
    """Search for files by name pattern and optionally by content."""
    try:
        p = Path(folder).expanduser().resolve()
        if not p.exists():
            return {"error": f"Folder not found: {folder}"}
        matches = []
        for item in p.rglob(pattern):
            if len(matches) >= 100:
                break
            match_info = {
                "name": item.name,
                "path": str(item),
                "type": "folder" if item.is_dir() else "file",
                "modified": datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            }
            if content_search and item.is_file():
                try:
                    text = item.read_text(encoding="utf-8", errors="replace")
                    if content_search.lower() in text.lower():
                        match_info["content_match"] = True
                    else:
                        continue
                except Exception:
                    continue
            matches.append(match_info)
        return {"folder": str(p), "pattern": pattern, "matches": matches, "count": len(matches)}
    except Exception as e:
        return {"error": str(e)}


def open_file(path: str) -> dict:
    """Open a file with its default application."""
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"File not found: {path}"}
        if os.name == "nt":
            os.startfile(str(p))
        else:
            subprocess.Popen(["xdg-open", str(p)])
        return {"success": True, "opened": str(p)}
    except Exception as e:
        return {"error": str(e)}


def open_folder(path: str) -> dict:
    """Open a folder in Windows Explorer."""
    try:
        p = Path(path).expanduser().resolve()
        if os.name == "nt":
            subprocess.Popen(["explorer", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])
        return {"success": True, "opened": str(p)}
    except Exception as e:
        return {"error": str(e)}


# ── Tool registry for the agent ────────────────────────────────────────────

TOOLS = {
    "list_directory": list_directory,
    "read_file": read_file,
    "write_file": write_file,
    "copy_file": copy_file,
    "move_file": move_file,
    "delete_file": delete_file,
    "create_folder": create_folder,
    "rename_file": rename_file,
    "search_files": search_files,
    "open_file": open_file,
    "open_folder": open_folder,
}

TOOL_SCHEMAS = [
    {
        "name": "list_directory",
        "description": "List files and folders at a given path. Use this to explore the file system.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Directory path to list"},
                "recursive": {"type": "boolean", "description": "List all subfolders too", "default": False},
            },
            "required": ["path"],
        },
    },
    {
        "name": "read_file",
        "description": "Read the contents of a text file, Word document, or Excel spreadsheet.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Full path to the file"},
                "max_chars": {"type": "integer", "description": "Max characters to read", "default": 8000},
            },
            "required": ["path"],
        },
    },
    {
        "name": "write_file",
        "description": "Write content to a file. Creates the file if it doesn't exist.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Full path to write to"},
                "content": {"type": "string", "description": "Content to write"},
                "overwrite": {"type": "boolean", "description": "Replace file if it exists", "default": False},
            },
            "required": ["path", "content"],
        },
    },
    {
        "name": "copy_file",
        "description": "Copy a file from one location to another.",
        "input_schema": {
            "type": "object",
            "properties": {
                "src": {"type": "string", "description": "Source file path"},
                "dst": {"type": "string", "description": "Destination path"},
            },
            "required": ["src", "dst"],
        },
    },
    {
        "name": "move_file",
        "description": "Move or rename a file to a new location.",
        "input_schema": {
            "type": "object",
            "properties": {
                "src": {"type": "string", "description": "Source file path"},
                "dst": {"type": "string", "description": "Destination path"},
            },
            "required": ["src", "dst"],
        },
    },
    {
        "name": "delete_file",
        "description": "Delete a file. Sends to Recycle Bin by default on Windows.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "File or folder path to delete"},
                "send_to_recycle": {"type": "boolean", "description": "Send to Recycle Bin instead of permanent delete", "default": True},
            },
            "required": ["path"],
        },
    },
    {
        "name": "create_folder",
        "description": "Create a new folder (and any parent folders needed).",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Folder path to create"},
            },
            "required": ["path"],
        },
    },
    {
        "name": "rename_file",
        "description": "Rename a file or folder.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Current path"},
                "new_name": {"type": "string", "description": "New filename (not full path, just the name)"},
            },
            "required": ["path", "new_name"],
        },
    },
    {
        "name": "search_files",
        "description": "Search for files by name pattern (e.g. '*.pdf') and optionally by content.",
        "input_schema": {
            "type": "object",
            "properties": {
                "folder": {"type": "string", "description": "Folder to search in"},
                "pattern": {"type": "string", "description": "Filename pattern, e.g. '*.docx', 'invoice*', '*report*'"},
                "content_search": {"type": "string", "description": "Search for this text inside files (optional)"},
            },
            "required": ["folder", "pattern"],
        },
    },
    {
        "name": "open_file",
        "description": "Open a file with its default application (e.g. Word, Excel, PDF viewer).",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "File path to open"},
            },
            "required": ["path"],
        },
    },
    {
        "name": "open_folder",
        "description": "Open a folder in Windows Explorer so the user can see it.",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Folder path to open"},
            },
            "required": ["path"],
        },
    },
]
